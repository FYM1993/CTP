from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HOLDINGS_HEADERS = [
    "记录ID",
    "合约代码",
    "品种名称",
    "方向",
    "手数",
    "开仓均价",
    "开仓日期",
    "推荐日期",
    "备注",
]

RECOMMENDATION_HEADERS = [
    "记录ID",
    "推荐日期",
    "推荐时间",
    "合约代码",
    "品种名称",
    "方向",
    "计划类型",
    "策略类型",
    "入场触发",
    "确认详情",
    "确认低点",
    "确认高点",
    "执行剧本",
    "市场阶段",
    "原始入场价",
    "原始第一止损位",
    "原始止损位",
    "原始第一止盈位",
    "原始止盈位",
    "原始第一止盈RR",
    "原始准入RR",
    "Phase2分",
    "原始交易故事摘要",
    "备注",
]


def default_holdings_root(project_root: Path | None = None) -> Path:
    root = project_root or Path(__file__).resolve().parents[2]
    return root / "docs" / "holdings"


def build_holdings_workbook_path(root_dir: Path, trade_date: str) -> Path:
    return root_dir / f"holdings-{trade_date}.xlsx"


def find_daily_holdings_workbook(*, root_dir: Path, trade_date: str) -> Path | None:
    path = build_holdings_workbook_path(root_dir, trade_date)
    return path if path.exists() else None


def ensure_daily_holdings_workbook(*, root_dir: Path, trade_date: str) -> Path:
    output = build_holdings_workbook_path(root_dir, trade_date)
    if output.exists():
        return output

    root_dir.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    hold_ws = workbook.active
    hold_ws.title = "持仓录入"
    hold_ws.append(HOLDINGS_HEADERS)
    rec_ws = workbook.create_sheet("原始推荐")
    rec_ws.append(RECOMMENDATION_HEADERS)
    _style_holdings_sheet(hold_ws)
    _style_recommendation_sheet(rec_ws)

    previous = _find_previous_workbook(root_dir, trade_date)
    if previous is not None:
        _copy_existing_rows(previous, workbook)

    workbook.save(output)
    return output


def _find_previous_workbook(root_dir: Path, trade_date: str) -> Path | None:
    target = datetime.strptime(trade_date, "%Y-%m-%d").date()
    candidates: list[tuple[object, Path]] = []
    for path in root_dir.glob("holdings-*.xlsx"):
        stamp = path.stem.removeprefix("holdings-")
        try:
            current_date = datetime.strptime(stamp, "%Y-%m-%d").date()
        except ValueError:
            continue
        if current_date < target:
            candidates.append((current_date, path))
    if not candidates:
        return None
    candidates.sort(key=lambda item: item[0])
    return candidates[-1][1]


def _copy_existing_rows(previous_path: Path, target_workbook: Workbook) -> None:
    previous = load_workbook(previous_path)
    for sheet_name in ("持仓录入", "原始推荐"):
        source = previous[sheet_name]
        target = target_workbook[sheet_name]
        source_headers = [cell.value for cell in source[1]]
        target_headers = [cell.value for cell in target[1]]
        source_index = {
            str(header): index
            for index, header in enumerate(source_headers)
            if header not in (None, "")
        }
        for row in source.iter_rows(min_row=2, values_only=True):
            if not any(value not in (None, "") for value in row):
                continue
            aligned = [
                row[source_index[header]] if header in source_index and source_index[header] < len(row) else None
                for header in target_headers
            ]
            target.append(aligned)


def _style_holdings_sheet(sheet) -> None:
    _style_header_row(sheet, HOLDINGS_HEADERS)
    sheet.freeze_panes = "A2"
    _set_widths(sheet, [14, 12, 12, 10, 8, 12, 12, 12, 28])
    date_format = "yyyy-mm-dd"
    for cell in sheet["G"][1:]:
        cell.number_format = date_format
    for cell in sheet["H"][1:]:
        cell.number_format = date_format
    _add_list_validation(sheet, "D2:D200", "做多,做空")


def _style_recommendation_sheet(sheet) -> None:
    _style_header_row(sheet, RECOMMENDATION_HEADERS)
    sheet.freeze_panes = "A2"
    _set_widths(
        sheet,
        [
            14,
            12,
            10,
            12,
            12,
            10,
            10,
            16,
            12,
            24,
            12,
            12,
            14,
            14,
            12,
            13,
            12,
            13,
            12,
            12,
            12,
            10,
            28,
            28,
        ],
    )
    for cell in sheet["B"][1:]:
        cell.number_format = "yyyy-mm-dd"
    for cell in sheet["C"][1:]:
        cell.number_format = "hh:mm"
    _add_list_validation(sheet, "F2:F200", "做多,做空")
    _add_list_validation(sheet, "G2:G200", "趋势,反转")


def _style_header_row(sheet, headers: list[str]) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    align = Alignment(horizontal="center", vertical="center")
    for index, _ in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=index)
        cell.fill = fill
        cell.font = font
        cell.alignment = align


def _set_widths(sheet, widths: list[float]) -> None:
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def _add_list_validation(sheet, cell_range: str, values: str) -> None:
    validation = DataValidation(type="list", formula1=f'"{values}"', allow_blank=True)
    validation.add(cell_range)
    sheet.add_data_validation(validation)

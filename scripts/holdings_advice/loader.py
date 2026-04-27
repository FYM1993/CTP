from __future__ import annotations

import json
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
import pandas as pd

from .workbook import HOLDINGS_HEADERS, RECOMMENDATION_HEADERS

_LONG_ALIASES = {"做多", "多", "long", "LONG", "Long"}
_SHORT_ALIASES = {"做空", "空", "short", "SHORT", "Short"}


def normalize_direction(value: str) -> str:
    text = str(value).strip()
    if text in _LONG_ALIASES:
        return "long"
    if text in _SHORT_ALIASES:
        return "short"
    raise ValueError(f"不支持的方向值: {value}")


def load_holding_contexts(
    path: Path,
    *,
    reports_dir: Path | None = None,
    autofill_recommendations: bool = True,
) -> list[dict]:
    if autofill_recommendations:
        _autofill_recommendation_sheet(path, reports_dir=reports_dir or _default_reports_dir())

    holdings_df = pd.read_excel(path, sheet_name="持仓录入")
    recommendation_df = pd.read_excel(path, sheet_name="原始推荐")

    recommendation_by_id: dict[str, dict] = {}
    for _, row in recommendation_df.iterrows():
        record_value = row.get("记录ID")
        if pd.isna(record_value):
            continue
        record_id = _base_record_id(record_value)
        recommendation_by_id[record_id] = {
            "record_id": record_id,
            "date": str(row.get("推荐日期") or ""),
            "time": str(row.get("推荐时间") or ""),
            "symbol": str(row.get("合约代码") or ""),
            "name": str(row.get("品种名称") or ""),
            "direction": normalize_direction(row.get("方向")),
            "entry_family": normalize_entry_family(row.get("计划类型")),
            "strategy_family": str(row.get("策略类型") or ""),
            "entry_signal_type": str(row.get("入场触发") or ""),
            "entry_signal_detail": str(row.get("确认详情") or ""),
            "confirmation_low": _as_float(row.get("确认低点")),
            "confirmation_high": _as_float(row.get("确认高点")),
            "selected_playbook": str(row.get("执行剧本") or ""),
            "market_stage": str(row.get("市场阶段") or ""),
            "entry": _as_float(row.get("原始入场价")),
            "first_stop": _as_float(row.get("原始第一止损位")),
            "stop": _as_float(row.get("原始止损位")),
            "tp1": _as_float(row.get("原始第一止盈位")),
            "tp2": _as_float(row.get("原始止盈位")),
            "rr": _as_float(row.get("原始第一止盈RR")),
            "admission_rr": _as_float(row.get("原始准入RR")),
            "score": _as_float(row.get("Phase2分")),
            "story_summary": str(row.get("原始交易故事摘要") or ""),
            "note": str(row.get("备注") or ""),
        }

    contexts: list[dict] = []
    for _, row in holdings_df.iterrows():
        record_value = row.get("记录ID")
        if pd.isna(record_value):
            continue
        record_id = _base_record_id(record_value)
        contexts.append(
            {
                "record_id": record_id,
                "holding": {
                    "record_id": record_id,
                    "symbol": str(row.get("合约代码") or ""),
                    "name": str(row.get("品种名称") or ""),
                    "direction": normalize_direction(row.get("方向")),
                    "size": _as_float(row.get("手数")),
                    "entry_price": _as_float(row.get("开仓均价")),
                    "open_date": str(row.get("开仓日期") or ""),
                    "recommendation_date": _date_text(row.get("推荐日期")),
                    "note": str(row.get("备注") or ""),
                },
                "recommendation": recommendation_by_id.get(record_id),
            }
        )
    return contexts


def normalize_entry_family(value: object) -> str:
    text = str(value or "").strip().lower()
    if text in {"趋势", "trend"}:
        return "trend"
    if text in {"反转", "reversal"}:
        return "reversal"
    return text


def _default_reports_dir() -> Path:
    return Path(__file__).resolve().parents[2] / "data" / "reports"


def _autofill_recommendation_sheet(path: Path, *, reports_dir: Path) -> None:
    workbook = load_workbook(path)
    changed = False
    if "持仓录入" not in workbook.sheetnames:
        return
    if "原始推荐" not in workbook.sheetnames:
        workbook.create_sheet("原始推荐")
        changed = True

    hold_sheet = workbook["持仓录入"]
    rec_sheet = workbook["原始推荐"]
    changed = _ensure_headers(hold_sheet, HOLDINGS_HEADERS) or changed
    changed = _ensure_headers(rec_sheet, RECOMMENDATION_HEADERS) or changed

    hold_headers = _header_index(hold_sheet)
    rec_headers = _header_index(rec_sheet)
    rec_rows = _recommendation_rows_by_id(rec_sheet, rec_headers)

    for row_idx in range(2, hold_sheet.max_row + 1):
        record_value = _cell_value(hold_sheet, row_idx, hold_headers, "记录ID")
        if _is_empty(record_value):
            continue
        record_id = _base_record_id(record_value)
        symbol = str(_cell_value(hold_sheet, row_idx, hold_headers, "合约代码") or "").strip()
        direction = _safe_normalize_direction(_cell_value(hold_sheet, row_idx, hold_headers, "方向"))
        if not symbol or direction not in {"long", "short"}:
            continue

        rec_row_idx = rec_rows.get(record_id)
        if rec_row_idx is None:
            rec_row_idx = rec_sheet.max_row + 1
            rec_rows[record_id] = rec_row_idx
            _set_cell(rec_sheet, rec_row_idx, rec_headers, "记录ID", f"{record_id}-推荐")
            changed = True

        recommendation_date = _date_text(
            _cell_value(hold_sheet, row_idx, hold_headers, "推荐日期")
            or _cell_value(rec_sheet, rec_row_idx, rec_headers, "推荐日期")
            or _cell_value(hold_sheet, row_idx, hold_headers, "开仓日期")
        )
        snapshot = _find_report_snapshot(
            reports_dir=reports_dir,
            recommendation_date=recommendation_date,
            symbol=symbol,
            direction=direction,
        )
        base_values = {
            "记录ID": f"{record_id}-推荐",
            "推荐日期": recommendation_date,
            "合约代码": symbol,
            "品种名称": _cell_value(hold_sheet, row_idx, hold_headers, "品种名称"),
            "方向": "做多" if direction == "long" else "做空",
        }
        if snapshot:
            base_values.update(snapshot)
        for header, value in base_values.items():
            if header not in rec_headers or _is_empty(value):
                continue
            if _is_empty(_cell_value(rec_sheet, rec_row_idx, rec_headers, header)):
                _set_cell(rec_sheet, rec_row_idx, rec_headers, header, value)
                changed = True

    if changed:
        workbook.save(path)


def _ensure_headers(sheet, headers: list[str]) -> bool:
    existing = [sheet.cell(row=1, column=index).value for index in range(1, sheet.max_column + 1)]
    changed = False
    if not any(value not in (None, "") for value in existing):
        for index, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=index).value = header
        return True
    existing_set = {str(value) for value in existing if value not in (None, "")}
    next_column = len(existing) + 1
    for header in headers:
        if header in existing_set:
            continue
        sheet.cell(row=1, column=next_column).value = header
        next_column += 1
        changed = True
    return changed


def _header_index(sheet) -> dict[str, int]:
    return {
        str(cell.value): index
        for index, cell in enumerate(sheet[1], start=1)
        if cell.value not in (None, "")
    }


def _recommendation_rows_by_id(sheet, headers: dict[str, int]) -> dict[str, int]:
    rows: dict[str, int] = {}
    for row_idx in range(2, sheet.max_row + 1):
        record_value = _cell_value(sheet, row_idx, headers, "记录ID")
        if _is_empty(record_value):
            continue
        rows[_base_record_id(record_value)] = row_idx
    return rows


def _cell_value(sheet, row_idx: int, headers: dict[str, int], header: str) -> Any:
    column = headers.get(header)
    if column is None:
        return None
    return sheet.cell(row=row_idx, column=column).value


def _set_cell(sheet, row_idx: int, headers: dict[str, int], header: str, value: Any) -> None:
    column = headers.get(header)
    if column is None:
        return
    sheet.cell(row=row_idx, column=column).value = value


def _find_report_snapshot(
    *,
    reports_dir: Path,
    recommendation_date: str,
    symbol: str,
    direction: str,
) -> dict[str, Any]:
    if not recommendation_date:
        return {}
    path = reports_dir / f"{recommendation_date}_targets.json"
    if not path.exists():
        return {}
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    rows = list(payload.get("targets") or []) + list(payload.get("watchlist") or [])
    for row in rows:
        if str(row.get("symbol") or "").strip().upper() != symbol.strip().upper():
            continue
        if _safe_normalize_direction(row.get("direction")) != direction:
            continue
        return _snapshot_to_recommendation_values(row)
    return {}


def _snapshot_to_recommendation_values(row: dict[str, Any]) -> dict[str, Any]:
    signal_bar = ((row.get("reversal_status") or {}).get("signal_bar") or {})
    return {
        "推荐时间": str(row.get("generated_time") or row.get("time") or ""),
        "合约代码": str(row.get("symbol") or ""),
        "品种名称": str(row.get("name") or ""),
        "方向": "做多" if _safe_normalize_direction(row.get("direction")) == "long" else "做空",
        "计划类型": _entry_family_cn(row.get("entry_family")),
        "策略类型": str(row.get("strategy_family") or ""),
        "入场触发": str(row.get("entry_signal_type") or ""),
        "确认详情": str(row.get("entry_signal_detail") or ""),
        "确认低点": _none_if_zero(_as_float(signal_bar.get("low"))),
        "确认高点": _none_if_zero(_as_float(signal_bar.get("high"))),
        "执行剧本": str(row.get("selected_playbook") or ""),
        "市场阶段": str(row.get("market_stage") or ""),
        "原始入场价": _none_if_zero(_as_float(row.get("entry"))),
        "原始第一止损位": _none_if_zero(_as_float(row.get("first_stop") or row.get("stop"))),
        "原始止损位": _none_if_zero(_as_float(row.get("stop"))),
        "原始第一止盈位": _none_if_zero(_as_float(row.get("tp1"))),
        "原始止盈位": _none_if_zero(_as_float(row.get("tp2"))),
        "原始第一止盈RR": _none_if_zero(_as_float(row.get("rr"))),
        "原始准入RR": _none_if_zero(_as_float(row.get("admission_rr"))),
        "Phase2分": _as_float(row.get("score")),
        "原始交易故事摘要": str(
            row.get("entry_signal_detail")
            or row.get("reason")
            or row.get("phase1_reason_summary")
            or row.get("reason_summary")
            or ""
        ),
    }


def _entry_family_cn(value: object) -> str:
    family = normalize_entry_family(value)
    if family == "trend":
        return "趋势"
    if family == "reversal":
        return "反转"
    return str(value or "")


def _safe_normalize_direction(value: object) -> str:
    try:
        return normalize_direction(str(value))
    except ValueError:
        return ""


def _date_text(value: object) -> str:
    if value is None or pd.isna(value):
        return ""
    if isinstance(value, (datetime, date, pd.Timestamp)):
        return pd.Timestamp(value).date().isoformat()
    text = str(value).strip()
    if not text:
        return ""
    try:
        return pd.Timestamp(text).date().isoformat()
    except (ValueError, TypeError):
        return text


def _is_empty(value: object) -> bool:
    if value is None:
        return True
    try:
        if pd.isna(value):
            return True
    except (TypeError, ValueError):
        pass
    return str(value).strip() == ""


def _none_if_zero(value: float) -> float | None:
    return None if value == 0.0 else value


def _base_record_id(value: object) -> str:
    return str(value).split("-")[0]


def _as_float(value: object) -> float:
    if pd.isna(value):
        return 0.0
    return float(value)

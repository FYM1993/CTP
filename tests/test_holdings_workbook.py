from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
SCRIPTS = ROOT / "scripts"
if str(SCRIPTS) not in sys.path:
    sys.path.insert(0, str(SCRIPTS))

from holdings_advice.workbook import (  # noqa: E402
    HOLDINGS_HEADERS,
    RECOMMENDATION_HEADERS,
    ensure_daily_holdings_workbook,
)


def test_ensure_daily_holdings_workbook_creates_schema(tmp_path: Path) -> None:
    output = ensure_daily_holdings_workbook(root_dir=tmp_path, trade_date="2026-04-23")

    assert output == tmp_path / "holdings-2026-04-23.xlsx"
    workbook = load_workbook(output)
    assert workbook.sheetnames == ["持仓录入", "原始推荐"]
    assert [cell.value for cell in workbook["持仓录入"][1]] == HOLDINGS_HEADERS
    assert [cell.value for cell in workbook["原始推荐"][1]] == RECOMMENDATION_HEADERS
    assert "推荐日期" in HOLDINGS_HEADERS
    assert "入场触发" in RECOMMENDATION_HEADERS
    assert "确认低点" in RECOMMENDATION_HEADERS
    assert "确认高点" in RECOMMENDATION_HEADERS


def test_ensure_daily_holdings_workbook_copies_previous_rows(tmp_path: Path) -> None:
    previous = ensure_daily_holdings_workbook(root_dir=tmp_path, trade_date="2026-04-22")
    workbook = load_workbook(previous)

    holdings_sheet = workbook["持仓录入"]
    recommendation_sheet = workbook["原始推荐"]
    holdings_sheet["A2"] = "H001-持仓"
    holdings_sheet["B2"] = "LH0"
    holdings_sheet["D2"] = "做多"
    holdings_sheet["E2"] = 2
    holdings_sheet["F2"] = 11480
    holdings_sheet["G2"] = "2026-04-22"
    holdings_sheet["H2"] = "旧备注"
    recommendation_sheet["A2"] = "H001-推荐"
    recommendation_sheet["B2"] = "2026-04-22"
    recommendation_sheet["C2"] = "09:12"
    recommendation_sheet["D2"] = "LH0"
    recommendation_sheet["F2"] = "做多"
    workbook.save(previous)

    current = ensure_daily_holdings_workbook(root_dir=tmp_path, trade_date="2026-04-23")

    current_workbook = load_workbook(current)
    assert current_workbook["持仓录入"]["A2"].value == "H001-持仓"
    assert current_workbook["持仓录入"]["B2"].value == "LH0"
    assert current_workbook["持仓录入"]["G2"].value == "2026-04-22"
    assert current_workbook["持仓录入"]["H2"].value == "旧备注"
    assert current_workbook["原始推荐"]["A2"].value == "H001-推荐"
    assert current_workbook["原始推荐"]["D2"].value == "LH0"

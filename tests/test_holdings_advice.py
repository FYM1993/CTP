from __future__ import annotations

import json
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parent.parent
SCRIPTS = ROOT / "scripts"
if str(SCRIPTS) not in sys.path:
    sys.path.insert(0, str(SCRIPTS))

from holdings_advice.analyzer import analyze_holding_record  # noqa: E402
from holdings_advice.loader import load_holding_contexts  # noqa: E402
from holdings_advice.presenter import (  # noqa: E402
    build_holdings_summary_line,
    format_holding_log_block,
    format_holding_terminal_alert,
)


def test_load_holding_contexts_pairs_rows_and_normalizes_direction(tmp_path: Path) -> None:
    workbook = Workbook()
    holdings_sheet = workbook.active
    holdings_sheet.title = "持仓录入"
    holdings_sheet.append(["记录ID", "合约代码", "品种名称", "方向", "手数", "开仓均价", "开仓日期", "备注"])
    holdings_sheet.append(["H001-持仓", "LH0", "生猪", "多", 2, 11480, "2026-04-22", ""])

    recommendation_sheet = workbook.create_sheet("原始推荐")
    recommendation_sheet.append(
        [
            "记录ID",
            "推荐日期",
            "推荐时间",
            "合约代码",
            "品种名称",
            "方向",
            "计划类型",
            "原始入场价",
            "原始第一止损位",
            "原始止损位",
            "原始第一止盈位",
            "原始止盈位",
            "原始第一止盈RR",
            "原始准入RR",
            "备注",
        ]
    )
    recommendation_sheet.append(
        ["H001-推荐", "2026-04-22", "09:12", "LH0", "生猪", "long", "趋势", 11490, 11160, 11142, 12265, 12445, 2.22, 2.74, ""]
    )
    path = tmp_path / "holdings-2026-04-23.xlsx"
    workbook.save(path)

    contexts = load_holding_contexts(path)

    assert len(contexts) == 1
    assert contexts[0]["record_id"] == "H001"
    assert contexts[0]["holding"]["direction"] == "long"
    assert contexts[0]["recommendation"]["direction"] == "long"


def test_load_holding_contexts_autofills_original_recommendation_from_json(tmp_path: Path) -> None:
    workbook = Workbook()
    holdings_sheet = workbook.active
    holdings_sheet.title = "持仓录入"
    holdings_sheet.append(["记录ID", "合约代码", "品种名称", "方向", "手数", "开仓均价", "开仓日期", "推荐日期", "备注"])
    holdings_sheet.append(["H001-持仓", "LH0", "生猪", "做多", 1, 11480, "2026-04-23", "2026-04-22", ""])

    recommendation_sheet = workbook.create_sheet("原始推荐")
    recommendation_sheet.append(
        [
            "记录ID",
            "推荐日期",
            "推荐时间",
            "合约代码",
            "品种名称",
            "方向",
            "计划类型",
            "策略类型",
            "入场触发",
            "确认详情",
            "确认低点",
            "确认高点",
            "执行剧本",
            "市场阶段",
            "原始入场价",
            "原始第一止损位",
            "原始止损位",
            "原始第一止盈位",
            "原始止盈位",
            "原始第一止盈RR",
            "原始准入RR",
            "Phase2分",
            "原始交易故事摘要",
            "备注",
        ]
    )
    path = tmp_path / "holdings-2026-04-23.xlsx"
    workbook.save(path)

    reports_dir = tmp_path / "reports"
    reports_dir.mkdir()
    (reports_dir / "2026-04-22_targets.json").write_text(
        json.dumps(
            {
                "date": "2026-04-22",
                "targets": [
                    {
                        "symbol": "LH0",
                        "name": "生猪",
                        "direction": "long",
                        "strategy_family": "trend_following",
                        "entry_family": "trend",
                        "entry_signal_type": "Pullback",
                        "entry_signal_detail": "markup 阶段顺势回踩",
                        "selected_playbook": "trend_pullback",
                        "market_stage": "clear_trend",
                        "entry": 11490.0,
                        "first_stop": 11160.0,
                        "stop": 11142.0,
                        "tp1": 12265.0,
                        "tp2": 12445.0,
                        "rr": 2.22,
                        "admission_rr": 2.74,
                        "score": 36.5,
                        "phase1_reason_summary": "价格/OI共振上行",
                        "reversal_status": {
                            "signal_bar": {"low": 11380.0, "high": 11600.0}
                        },
                    }
                ],
                "watchlist": [],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    contexts = load_holding_contexts(path, reports_dir=reports_dir)

    assert len(contexts) == 1
    recommendation = contexts[0]["recommendation"]
    assert recommendation["entry_family"] == "trend"
    assert recommendation["strategy_family"] == "trend_following"
    assert recommendation["entry_signal_type"] == "Pullback"
    assert recommendation["entry_signal_detail"] == "markup 阶段顺势回踩"
    assert recommendation["confirmation_low"] == 11380.0
    assert recommendation["selected_playbook"] == "trend_pullback"
    assert recommendation["score"] == 36.5

    saved = load_workbook(path, data_only=True)
    saved_sheet = saved["原始推荐"]
    saved_headers = [cell.value for cell in saved_sheet[1]]
    saved_row = {header: saved_sheet.cell(row=2, column=index + 1).value for index, header in enumerate(saved_headers)}
    assert saved_row["记录ID"] == "H001-推荐"
    assert saved_row["入场触发"] == "Pullback"
    assert saved_row["确认低点"] == 11380.0
    assert saved_row["原始入场价"] == 11490.0


def test_load_holding_contexts_keeps_manual_recommendation_over_json(tmp_path: Path) -> None:
    workbook = Workbook()
    holdings_sheet = workbook.active
    holdings_sheet.title = "持仓录入"
    holdings_sheet.append(["记录ID", "合约代码", "品种名称", "方向", "手数", "开仓均价", "开仓日期", "推荐日期", "备注"])
    holdings_sheet.append(["H001-持仓", "LH0", "生猪", "做多", 1, 11480, "2026-04-23", "2026-04-22", ""])

    recommendation_sheet = workbook.create_sheet("原始推荐")
    recommendation_sheet.append(
        [
            "记录ID",
            "推荐日期",
            "推荐时间",
            "合约代码",
            "品种名称",
            "方向",
            "计划类型",
            "策略类型",
            "入场触发",
            "确认详情",
            "确认低点",
            "确认高点",
            "执行剧本",
            "市场阶段",
            "原始入场价",
            "原始第一止损位",
            "原始止损位",
            "原始第一止盈位",
            "原始止盈位",
            "原始第一止盈RR",
            "原始准入RR",
            "Phase2分",
            "原始交易故事摘要",
            "备注",
        ]
    )
    recommendation_sheet.append(
        [
            "H001-推荐",
            "2026-04-22",
            "",
            "LH0",
            "生猪",
            "做多",
            "趋势",
            "",
            "ManualSignal",
            "用户手填确认",
            11200.0,
            "",
            "",
            "",
            11500.0,
            "",
            11100.0,
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ]
    )
    path = tmp_path / "holdings-2026-04-23.xlsx"
    workbook.save(path)

    reports_dir = tmp_path / "reports"
    reports_dir.mkdir()
    (reports_dir / "2026-04-22_targets.json").write_text(
        json.dumps(
            {
                "date": "2026-04-22",
                "targets": [
                    {
                        "symbol": "LH0",
                        "name": "生猪",
                        "direction": "long",
                        "entry_family": "trend",
                        "entry_signal_type": "Pullback",
                        "entry_signal_detail": "JSON确认",
                        "entry": 11490.0,
                        "stop": 11142.0,
                    }
                ],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    contexts = load_holding_contexts(path, reports_dir=reports_dir)

    recommendation = contexts[0]["recommendation"]
    assert recommendation["entry_signal_type"] == "ManualSignal"
    assert recommendation["entry_signal_detail"] == "用户手填确认"
    assert recommendation["confirmation_low"] == 11200.0
    assert recommendation["entry"] == 11500.0
    assert recommendation["stop"] == 11100.0


def test_analyze_holding_record_marks_raise_stop_when_current_stop_improves() -> None:
    result = analyze_holding_record(
        holding={
            "record_id": "H001",
            "symbol": "LH0",
            "name": "生猪",
            "direction": "long",
            "size": 2,
            "entry_price": 11480.0,
        },
        recommendation={
            "entry_family": "trend",
            "entry": 11490.0,
            "first_stop": 11160.0,
            "stop": 11142.0,
            "tp1": 12265.0,
            "tp2": 12445.0,
            "rr": 2.22,
            "admission_rr": 2.74,
        },
        current_plan={
            "entry": 11545.0,
            "first_stop": 11165.0,
            "stop": 11151.0,
            "tp1": 12265.0,
            "tp2": 12445.0,
            "rr": 1.83,
            "admission_rr": 2.28,
            "actionable": True,
        },
        hold_eval={"hold_valid": True, "story_family": "trend"},
        current_price=11540.0,
        minimum_tick=1.0,
    )

    assert result["action"] == "继续持有"
    assert result["management_action"] == "上移止损"
    assert "原有顺势逻辑仍成立" in result["reason"]
    assert "11142上移至11151" in result["reason"]


def test_analyze_holding_record_marks_exit_when_trade_story_invalidates() -> None:
    result = analyze_holding_record(
        holding={
            "record_id": "H002",
            "symbol": "PS0",
            "name": "多晶硅",
            "direction": "short",
            "size": 1,
            "entry_price": 199.0,
        },
        recommendation={
            "entry_family": "trend",
            "entry": 199.0,
            "first_stop": 205.0,
            "stop": 205.0,
            "tp1": 192.0,
            "tp2": 189.0,
            "rr": 1.16,
            "admission_rr": 1.67,
        },
        current_plan={
            "entry": 198.2,
            "first_stop": 204.0,
            "stop": 204.0,
            "tp1": 192.0,
            "tp2": 189.0,
            "rr": 0.6,
            "admission_rr": 0.9,
        },
        hold_eval={"hold_valid": False, "story_family": "trend"},
        current_price=201.0,
        minimum_tick=1.0,
    )

    assert result["action"] == "平仓"
    assert "原有顺势逻辑已失效" in result["reason"]


def test_analyze_holding_record_exits_when_reversal_confirmation_level_breaks() -> None:
    result = analyze_holding_record(
        holding={
            "record_id": "H003",
            "symbol": "LH0",
            "name": "生猪",
            "direction": "long",
            "size": 1,
            "entry_price": 11450.0,
        },
        recommendation={
            "entry_family": "reversal",
            "entry_signal_type": "Spring",
            "confirmation_low": 11200.0,
            "entry": 11450.0,
            "first_stop": 11050.0,
            "stop": 11050.0,
            "tp1": 12265.0,
            "tp2": 12445.0,
            "rr": 2.70,
            "admission_rr": 3.29,
        },
        current_plan={
            "entry": 11190.0,
            "first_stop": 11050.0,
            "stop": 11050.0,
            "tp1": 12265.0,
            "tp2": 12445.0,
            "rr": 5.0,
            "admission_rr": 6.0,
        },
        hold_eval={"hold_valid": True, "story_family": "reversal"},
        current_price=11190.0,
        minimum_tick=1.0,
    )

    assert result["action"] == "平仓"
    assert "原反转确认位11200已失守" in result["reason"]


def test_analyze_holding_record_exits_when_trend_pullback_confirmation_level_breaks() -> None:
    result = analyze_holding_record(
        holding={
            "record_id": "H004",
            "symbol": "LH0",
            "name": "生猪",
            "direction": "long",
            "size": 1,
            "entry_price": 11450.0,
        },
        recommendation={
            "entry_family": "trend",
            "entry_signal_type": "Pullback",
            "confirmation_low": 11380.0,
            "entry": 11450.0,
            "first_stop": 11128.0,
            "stop": 11128.0,
            "tp1": 12265.0,
            "tp2": 12445.0,
            "rr": 2.70,
            "admission_rr": 3.29,
        },
        current_plan={
            "entry": 11370.0,
            "first_stop": 11128.0,
            "stop": 11128.0,
            "tp1": 12265.0,
            "tp2": 12445.0,
            "rr": 6.0,
            "admission_rr": 7.0,
        },
        hold_eval={"hold_valid": True, "story_family": "trend"},
        current_price=11370.0,
        minimum_tick=1.0,
    )

    assert result["action"] == "平仓"
    assert "原趋势回踩确认位11380已失守" in result["reason"]


def test_analyze_holding_record_uses_conservative_wording_when_story_snapshot_missing() -> None:
    result = analyze_holding_record(
        holding={
            "record_id": "H005",
            "symbol": "LH0",
            "name": "生猪",
            "direction": "long",
            "size": 1,
            "entry_price": 11450.0,
        },
        recommendation={
            "entry": 11450.0,
            "first_stop": 11128.0,
            "stop": 11128.0,
            "tp1": 12265.0,
            "tp2": 12445.0,
            "rr": 2.70,
            "admission_rr": 3.29,
        },
        current_plan={
            "entry": 11405.0,
            "first_stop": 11128.0,
            "stop": 11128.0,
            "tp1": 12265.0,
            "tp2": 12445.0,
            "rr": 2.70,
            "admission_rr": 3.29,
        },
        hold_eval={"hold_valid": True, "snapshot_missing": True},
        current_price=11405.0,
        minimum_tick=1.0,
    )

    assert result["action"] == "继续持有"
    assert "缺少原始交易剧本" in result["reason"]


def test_analyze_holding_record_reduces_when_current_stop_risk_exceeds_account_cap() -> None:
    result = analyze_holding_record(
        holding={
            "record_id": "H006",
            "symbol": "LH0",
            "name": "生猪",
            "direction": "long",
            "size": 4,
            "entry_price": 11450.0,
        },
        recommendation={
            "entry_family": "trend",
            "entry": 11450.0,
            "first_stop": 11100.0,
            "stop": 11100.0,
            "tp1": 12200.0,
            "tp2": 12400.0,
            "rr": 2.70,
            "admission_rr": 3.29,
        },
        current_plan={
            "entry": 11400.0,
            "first_stop": 11100.0,
            "stop": 11100.0,
            "tp1": 12200.0,
            "tp2": 12400.0,
            "rr": 2.70,
            "admission_rr": 3.29,
        },
        hold_eval={"hold_valid": True, "story_family": "trend"},
        current_price=11400.0,
        minimum_tick=1.0,
        account_equity=1_000_000.0,
        risk_per_trade_pct=0.015,
        contract_multiplier=16.0,
    )

    assert result["action"] == "减仓观察"
    assert result["management_action"] == "减仓"
    assert result["risk_control"]["risk_budget"] == pytest.approx(15_000.0)
    assert result["risk_control"]["max_safe_size"] == 3
    assert "当前到止损" in result["reason"]


def test_analyze_holding_record_marks_exit_when_long_position_hits_original_stop() -> None:
    result = analyze_holding_record(
        holding={
            "record_id": "H004",
            "symbol": "PS0",
            "name": "多晶硅",
            "direction": "long",
            "size": 1,
            "entry_price": 43125.0,
        },
        recommendation={
            "entry_family": "reversal",
            "entry": 43125.0,
            "first_stop": 40988.0,
            "stop": 40988.0,
            "tp1": 50175.0,
            "tp2": 52842.0,
            "rr": 3.30,
            "admission_rr": 4.55,
        },
        current_plan={
            "entry": 40980.0,
            "first_stop": 40988.0,
            "stop": 40988.0,
            "tp1": 50175.0,
            "tp2": 52842.0,
            "rr": 4.7,
            "admission_rr": 5.9,
        },
        hold_eval={"hold_valid": True, "story_family": "reversal"},
        current_price=40980.0,
        minimum_tick=1.0,
    )

    assert result["action"] == "平仓"
    assert "触及止损40988" in result["reason"]


def test_analyze_holding_record_marks_exit_when_short_position_hits_original_stop() -> None:
    result = analyze_holding_record(
        holding={
            "record_id": "H005",
            "symbol": "PS0",
            "name": "多晶硅",
            "direction": "short",
            "size": 1,
            "entry_price": 43125.0,
        },
        recommendation={
            "entry_family": "trend",
            "entry": 43125.0,
            "first_stop": 44200.0,
            "stop": 44200.0,
            "tp1": 41000.0,
            "tp2": 39800.0,
            "rr": 1.98,
            "admission_rr": 3.09,
        },
        current_plan={
            "entry": 44210.0,
            "first_stop": 44200.0,
            "stop": 44200.0,
            "tp1": 41000.0,
            "tp2": 39800.0,
            "rr": 2.99,
            "admission_rr": 4.10,
        },
        hold_eval={"hold_valid": True, "story_family": "trend"},
        current_price=44210.0,
        minimum_tick=1.0,
    )

    assert result["action"] == "平仓"
    assert "触及止损44200" in result["reason"]


def test_analyze_holding_record_describes_reversal_transition_to_trend_followthrough() -> None:
    result = analyze_holding_record(
        holding={
            "record_id": "H003",
            "symbol": "LH0",
            "name": "生猪",
            "direction": "long",
            "size": 1,
            "entry_price": 11450.0,
        },
        recommendation={
            "entry_family": "reversal",
            "entry": 11450.0,
            "first_stop": 11128.0,
            "stop": 11128.0,
            "tp1": 12265.0,
            "tp2": 12445.0,
            "rr": 2.70,
            "admission_rr": 3.29,
        },
        current_plan={
            "entry": 11520.0,
            "first_stop": 11141.0,
            "stop": 11141.0,
            "tp1": 12265.0,
            "tp2": 12445.0,
            "rr": 2.30,
            "admission_rr": 2.90,
        },
        hold_eval={"hold_valid": True, "story_family": "reversal", "story_transition": "trend_followthrough"},
        current_price=11520.0,
        minimum_tick=1.0,
    )

    assert result["action"] == "继续持有"
    assert result["management_action"] == "上移止损"
    assert "原有反转逻辑已进入顺势延续阶段" in result["reason"]


def test_format_holding_log_block_includes_original_and_current_plan() -> None:
    text = format_holding_log_block(
        {
            "record_id": "H001",
            "name": "生猪",
            "direction_cn": "做多",
            "size": 2,
            "entry_price": 11480.0,
            "current_price": 11540.0,
            "original_plan_line": "原始计划: 入场11490 第一止损11160 止损11142 第一止盈12265 止盈12445 第一止盈RR 2.22 准入RR 2.74",
            "current_plan_line": "当前重评: 入场11545 第一止损11165 止损11151 第一止盈12265 止盈12445 第一止盈RR 1.83 准入RR 2.28",
            "action": "继续持有",
            "reason": "原始趋势未破坏",
        }
    )

    assert "H001 生猪 做多 2手 开仓均价11480 当前价11540" in text
    assert "原始计划:" in text
    assert "当前重评:" in text
    assert "持仓建议: 继续持有" in text


def test_format_holding_log_block_describes_raise_stop_as_hold_management() -> None:
    text = format_holding_log_block(
        {
            "record_id": "H001",
            "name": "生猪",
            "direction_cn": "做多",
            "size": 1,
            "entry_price": 11450.0,
            "current_price": 11475.0,
            "action": "继续持有",
            "management_action": "上移止损",
            "reason": "原有顺势逻辑仍成立，保护止损可从11128上移至11136",
            "original_plan_line": "原始计划: 入场11450 第一止损11128 止损11128 第一止盈12265 止盈12445 第一止盈RR 2.70 准入RR 3.29",
            "current_plan_line": "当前重评: 入场11475 第一止损11135.82 止损11135.82 第一止盈12265.24 止盈12445 第一止盈RR 2.71 准入RR 3.31",
            "original_plan": {"stop": 11128.0},
            "current_plan": {"stop": 11135.82, "tp1": 12265.24},
        }
    )

    assert "持仓建议: 继续持有（保护止损上移）" in text
    assert "当前状态: 原有顺势逻辑仍成立，保护止损已抬高到11135.82，第一止盈继续看12265.24" in text
    assert "建议原因: 原有顺势逻辑仍成立，保护止损可从11128上移至11136" in text


def test_format_holding_status_text_keeps_story_before_raise_stop() -> None:
    from holdings_advice.presenter import format_holding_status_text

    result = {
        "action": "继续持有",
        "management_action": "上移止损",
        "reason": "原有顺势逻辑仍成立，保护止损可从11128上移至11141",
        "current_plan": {"stop": 11141.0, "tp1": 12265.0},
    }

    text = format_holding_status_text(result)

    assert text.startswith("原有顺势逻辑仍成立")
    assert "保护止损已抬高到11141" in text


def test_format_holding_terminal_alert_describes_raise_stop_as_hold_management() -> None:
    text = format_holding_terminal_alert(
        {
            "record_id": "H001",
            "name": "生猪",
            "action": "继续持有",
            "management_action": "上移止损",
            "reason": "原有顺势逻辑仍成立，保护止损可从11128上移至11141",
            "current_price": 11500.0,
            "original_plan": {
                "stop": 11128.0,
            },
            "current_plan": {
                "stop": 11141.0,
                "tp1": 12265.0,
            },
        }
    )

    assert "继续持有（原有顺势逻辑仍成立；保护止损上移 11128 -> 11141）" in text
    assert "当前价 11500" in text


def test_build_holdings_summary_line_treats_raise_stop_as_hold_subtype() -> None:
    text = build_holdings_summary_line(
        [
            {"action": "继续持有", "management_action": "上移止损"},
        ]
    )

    assert "继续持有1（其中保护止损上移1）" in text
    assert "减仓观察0" in text
    assert "平仓0" in text
    assert "计划失效" not in text


def test_format_holding_terminal_alert_includes_exit_reason() -> None:
    text = format_holding_terminal_alert(
        {
            "record_id": "H002",
            "name": "多晶硅",
            "action": "平仓",
            "reason": "原有顺势逻辑已失效，建议退出",
            "current_price": 201.0,
            "current_plan": {
                "stop": 204.0,
                "tp1": 192.0,
            },
        }
    )

    assert "平仓（原有顺势逻辑已失效）" in text
    assert "当前价 201" in text


def test_should_emit_terminal_alert_reemits_when_protection_level_changes() -> None:
    first = {
        "record_id": "H001",
        "action": "继续持有",
        "management_action": "上移止损",
        "current_plan": {"stop": 11141.0},
    }
    second_same = {
        "record_id": "H001",
        "action": "继续持有",
        "management_action": "上移止损",
        "current_plan": {"stop": 11141.0},
    }
    third_changed = {
        "record_id": "H001",
        "action": "继续持有",
        "management_action": "上移止损",
        "current_plan": {"stop": 11155.0},
    }

    from holdings_advice.presenter import should_emit_terminal_alert  # noqa: E402

    assert should_emit_terminal_alert(first, {}) is True
    assert should_emit_terminal_alert(second_same, {"H001": "继续持有|上移止损|11141"}) is False
    assert should_emit_terminal_alert(third_changed, {"H001": "继续持有|上移止损|11141"}) is True


def test_format_holding_log_block_uses_exit_for_invalidated_trade_story() -> None:
    text = format_holding_log_block(
        {
            "record_id": "H002",
            "name": "多晶硅",
            "direction_cn": "做空",
            "size": 1,
            "entry_price": 199.0,
            "current_price": 201.0,
            "original_plan_line": "原始计划: 入场199 止损205 第一止盈192 止盈189 第一止盈RR 1.16 准入RR 1.67",
            "current_plan_line": "当前重评: 入场198.20 止损204 第一止盈192 止盈189 第一止盈RR 0.60 准入RR 0.90",
            "action": "平仓",
            "reason": "原有顺势逻辑已失效，建议退出",
        }
    )

    assert "持仓建议: 平仓" in text
    assert "当前状态: 原有交易故事已失效，优先执行退出" in text
    assert "建议原因: 原有顺势逻辑已失效，建议退出" in text


def test_format_holding_log_block_uses_trim_state_when_first_target_reached() -> None:
    text = format_holding_log_block(
        {
            "record_id": "H004",
            "name": "尿素",
            "direction_cn": "做多",
            "size": 1,
            "entry_price": 2034.0,
            "current_price": 2052.0,
            "original_plan_line": "原始计划: 入场2034 第一止损2006 止损2006 第一止盈2051 止盈2058 第一止盈RR 0.61 准入RR 0.86",
            "current_plan_line": "当前重评: 入场2034 第一止损2012 止损2012 第一止盈2062 止盈2068 第一止盈RR 1.27 准入RR 1.55",
            "action": "减仓观察",
            "reason": "第一止盈位已达到或接近，建议管理仓位",
            "current_plan": {
                "stop": 2012.0,
                "tp1": 2062.0,
            },
        }
    )

    assert "持仓建议: 减仓观察" in text
    assert "当前状态: 已进入第一兑现区，优先执行减仓管理" in text
